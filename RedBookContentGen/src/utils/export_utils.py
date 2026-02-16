#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出工具模块

提供批量结果导出功能，支持 Excel 和 ZIP 打包
"""

import io
import zipfile
from pathlib import Path
from typing import Dict, List
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from src.core.logger import Logger


class ExportUtils:
    """导出工具类"""
    
    @staticmethod
    def export_batch_to_excel(batch_result: Dict) -> bytes:
        """
        将批量生成结果导出为 Excel 文件
        
        Args:
            batch_result: 批量生成结果字典
            
        Returns:
            Excel 文件的字节数据
            
        Raises:
            ImportError: 如果 openpyxl 未安装
        """
        if not HAS_OPENPYXL:
            raise ImportError("需要安装 openpyxl 库才能导出 Excel 文件")
        
        Logger.info(
            "开始导出批量结果到 Excel",
            logger_name="export_utils",
            batch_id=batch_result.get("batch_id"),
        )
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "批量生成结果"
        
        # 设置标题样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入标题行
        headers = ["序号", "状态", "输入文本", "标题", "内容", "标签", "错误信息"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 40
        
        # 写入数据行
        results = batch_result.get("results", [])
        for i, result in enumerate(results, start=2):
            ws.cell(row=i, column=1, value=result.get("index", 0) + 1)
            ws.cell(row=i, column=2, value="成功" if result.get("status") == "success" else "失败")
            ws.cell(row=i, column=3, value=result.get("input_text", ""))
            
            # 提取数据
            data = result.get("data")
            if data:
                titles = data.get("titles", [])
                ws.cell(row=i, column=4, value=titles[0] if titles else "")
                ws.cell(row=i, column=5, value=data.get("content", ""))
                ws.cell(row=i, column=6, value=", ".join(data.get("tags", [])))
            else:
                ws.cell(row=i, column=4, value="")
                ws.cell(row=i, column=5, value="")
                ws.cell(row=i, column=6, value="")
            
            ws.cell(row=i, column=7, value=result.get("error", ""))
        
        # 添加汇总信息
        summary = batch_result.get("summary", {})
        summary_row = len(results) + 3
        ws.cell(row=summary_row, column=1, value="汇总信息")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value=f"总数: {summary.get('total', 0)}")
        ws.cell(row=summary_row + 2, column=1, value=f"成功: {summary.get('success', 0)}")
        ws.cell(row=summary_row + 3, column=1, value=f"失败: {summary.get('failed', 0)}")
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        Logger.info(
            "Excel 导出完成",
            logger_name="export_utils",
            batch_id=batch_result.get("batch_id"),
            total_rows=len(results),
        )
        
        return output.getvalue()
    
    @staticmethod
    def create_batch_zip(batch_result: Dict, output_dir: Path) -> bytes:
        """
        将批量生成结果打包为 ZIP 文件
        
        包含：
        - Excel 汇总文件
        - 所有生成的图片文件
        - 批次信息文本文件
        
        Args:
            batch_result: 批量生成结果字典
            output_dir: 输出目录路径
            
        Returns:
            ZIP 文件的字节数据
        """
        Logger.info(
            "开始创建批量结果 ZIP 包",
            logger_name="export_utils",
            batch_id=batch_result.get("batch_id"),
        )
        
        # 创建内存中的 ZIP 文件
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 1. 添加 Excel 汇总文件
            if HAS_OPENPYXL:
                try:
                    excel_data = ExportUtils.export_batch_to_excel(batch_result)
                    batch_id = batch_result.get("batch_id", "unknown")
                    zip_file.writestr(f"{batch_id}_summary.xlsx", excel_data)
                except Exception as e:
                    Logger.warning(
                        "添加 Excel 文件失败",
                        logger_name="export_utils",
                        error=str(e),
                    )
            
            # 2. 添加批次信息文本文件
            batch_info = ExportUtils._create_batch_info_text(batch_result)
            zip_file.writestr("batch_info.txt", batch_info)
            
            # 3. 添加生成的图片文件（如果有）
            results = batch_result.get("results", [])
            for result in results:
                if result.get("status") == "success":
                    data = result.get("data", {})
                    timestamp = data.get("timestamp", "")
                    
                    # 查找该批次的图片文件
                    if timestamp:
                        image_pattern = f"*{timestamp}*.png"
                        for image_file in output_dir.glob(image_pattern):
                            # 添加到 ZIP，使用相对路径
                            arcname = f"images/{image_file.name}"
                            zip_file.write(image_file, arcname=arcname)
        
        zip_buffer.seek(0)
        
        Logger.info(
            "ZIP 包创建完成",
            logger_name="export_utils",
            batch_id=batch_result.get("batch_id"),
        )
        
        return zip_buffer.getvalue()
    
    @staticmethod
    def _create_batch_info_text(batch_result: Dict) -> str:
        """
        创建批次信息文本
        
        Args:
            batch_result: 批量生成结果字典
            
        Returns:
            批次信息文本
        """
        batch_id = batch_result.get("batch_id", "unknown")
        summary = batch_result.get("summary", {})
        results = batch_result.get("results", [])
        
        lines = [
            "=" * 60,
            "批量生成结果汇总",
            "=" * 60,
            f"批次ID: {batch_id}",
            f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"总任务数: {summary.get('total', 0)}",
            f"成功数量: {summary.get('success', 0)}",
            f"失败数量: {summary.get('failed', 0)}",
            "",
            "=" * 60,
            "详细结果",
            "=" * 60,
            "",
        ]
        
        for result in results:
            index = result.get("index", 0) + 1
            status = "✓ 成功" if result.get("status") == "success" else "✗ 失败"
            input_text = result.get("input_text", "")
            
            lines.append(f"{index}. {status}")
            lines.append(f"   输入: {input_text}")
            
            if result.get("status") == "success":
                data = result.get("data", {})
                titles = data.get("titles", [])
                if titles:
                    lines.append(f"   标题: {titles[0]}")
                tags = data.get("tags", [])
                if tags:
                    lines.append(f"   标签: {', '.join(tags)}")
            else:
                error = result.get("error", "未知错误")
                lines.append(f"   错误: {error}")
            
            lines.append("")
        
        return "\n".join(lines)
