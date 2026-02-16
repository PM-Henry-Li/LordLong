#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
老北京文化·小红书内容生成器
读取文档内容，生成小红书文案和AI绘画提示词，保存到Excel和文件夹
"""

import os
import json
from datetime import datetime
import re
from typing import List, Dict, Tuple, Optional, TYPE_CHECKING, Any, Callable

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import openai

from src.core.logger import Logger
from src.core.api_handler import APIHandler
from src.core.retry_handler import ErrorHandler
from src.core.exceptions import (
    FileNotFoundError as CustomFileNotFoundError,
    ContentValidationError,
    ContentSafetyError,
    wrap_exception,
)

if TYPE_CHECKING:
    from src.core.config_manager import ConfigManager


class RedBookContentGenerator:
    """小红书内容生成器"""

    def __init__(self, config_manager: Optional["ConfigManager"] = None, config_path: str = "config/config.json") -> None:
        """
        初始化生成器

        Args:
            config_manager: ConfigManager 实例（推荐使用）
            config_path: 配置文件路径（向后兼容，当 config_manager 为 None 时使用）
        """
        # 支持两种初始化方式：新方式（ConfigManager）和旧方式（config_path）
        if config_manager is not None:
            self.config_manager = config_manager
            self._use_config_manager = True
        else:
            # 向后兼容：如果没有传入 ConfigManager，则使用旧的配置加载方式
            from src.core.config_manager import ConfigManager

            self.config_manager = ConfigManager(config_path)
            self._use_config_manager = True

        # 初始化日志系统
        Logger.initialize(self.config_manager)
        self.logger = Logger.get_logger("content_generator")

        # 初始化缓存管理器
        self._init_cache()

        # 初始化速率限制器
        self._init_rate_limiter()

        # 初始化 API 处理器
        self.api_handler = APIHandler(
            rpm_limiter=self.rpm_limiter if hasattr(self, "rpm_limiter") else None,
            tpm_limiter=self.tpm_limiter if hasattr(self, "tpm_limiter") else None,
            rate_limit_enabled=self._rate_limit_enabled if hasattr(self, "_rate_limit_enabled") else False,
            logger_name="content_generator",
        )

        self.setup_paths()

        # API Key会在调用时检查，这里不需要初始化

    def _init_cache(self) -> None:
        """初始化缓存管理器"""
        from src.core.cache_manager import CacheManager

        # 检查是否启用缓存
        cache_enabled = self.config_manager.get("cache.enabled", True)

        if cache_enabled:
            # 获取缓存配置
            cache_ttl = self.config_manager.get("cache.ttl", 3600)  # 默认1小时
            cache_max_size = self.config_manager.get("cache.max_size", 1000)  # 默认1000条

            # 确保 max_size 是整数（处理可能的字符串配置）
            if isinstance(cache_max_size, str):
                # 如果是字符串（如 "1GB"），使用默认值
                cache_max_size = 1000

            # 创建缓存管理器实例
            self.cache = CacheManager(max_size=cache_max_size, default_ttl=cache_ttl)
            self._cache_enabled = True

            Logger.info("缓存已启用", logger_name="content_generator", ttl=cache_ttl, max_size=cache_max_size)
        else:
            self.cache = None
            self._cache_enabled = False
            Logger.info("缓存已禁用", logger_name="content_generator")

    def _init_rate_limiter(self) -> None:
        """初始化速率限制器"""
        from src.core.rate_limiter import RateLimiter

        # 检查是否启用速率限制
        rate_limit_enabled = self.config_manager.get("rate_limit.openai.enable_rate_limit", True)

        if rate_limit_enabled:
            # 获取速率限制配置
            rpm = self.config_manager.get("rate_limit.openai.requests_per_minute", 60)
            tpm = self.config_manager.get("rate_limit.openai.tokens_per_minute", 90000)

            # 创建速率限制器
            # RPM 限制器：每分钟请求数
            self.rpm_limiter = RateLimiter(rate=rpm / 60.0, capacity=rpm)

            # TPM 限制器：每分钟 token 数
            self.tpm_limiter = RateLimiter(rate=tpm / 60.0, capacity=tpm)

            self._rate_limit_enabled = True

            Logger.info(
                "速率限制已启用", logger_name="content_generator", requests_per_minute=rpm, tokens_per_minute=tpm
            )
        else:
            self.rpm_limiter = None
            self.tpm_limiter = None
            self._rate_limit_enabled = False

            Logger.info("速率限制已禁用", logger_name="content_generator")

    def _generate_cache_key(self, raw_content: str) -> str:
        """
        生成缓存键（基于输入内容的hash）

        Args:
            raw_content: 原始输入内容

        Returns:
            缓存键（SHA256哈希值）
        """
        import hashlib

        # 使用输入内容的hash作为缓存键
        content_hash = hashlib.sha256(raw_content.encode("utf - 8")).hexdigest()

        # 添加前缀以区分不同类型的缓存
        return f"content_gen:{content_hash}"

    def get_cache_stats(self) -> Optional[Dict[str, Any]]:
        """
        获取缓存统计信息

        Returns:
            缓存统计字典，如果缓存未启用则返回 None
        """
        if self._cache_enabled and self.cache is not None:
            return self.cache.get_stats()
        return None

    def get_rate_limit_stats(self) -> Optional[Dict[str, Any]]:
        """
        获取速率限制统计信息

        Returns:
            速率限制统计字典，如果速率限制未启用则返回 None
        """
        if not self._rate_limit_enabled:
            return None

        stats: Dict[str, Any] = {
            "enabled": True,
            "rpm": {
                "available_tokens": self.rpm_limiter.get_available_tokens() if self.rpm_limiter else None,
                "capacity": self.rpm_limiter.get_capacity() if self.rpm_limiter else None,
                "rate": self.rpm_limiter.get_rate() if self.rpm_limiter else None,
            },
            "tpm": {
                "available_tokens": self.tpm_limiter.get_available_tokens() if self.tpm_limiter else None,
                "capacity": self.tpm_limiter.get_capacity() if self.tpm_limiter else None,
                "rate": self.tpm_limiter.get_rate() if self.tpm_limiter else None,
            },
        }

        return stats

    def clear_cache(self) -> None:
        """清空缓存"""
        if self._cache_enabled and self.cache is not None:
            self.cache.clear()
            Logger.info("缓存已清空", logger_name="content_generator")
        else:
            Logger.warning("缓存未启用，无需清空", logger_name="content_generator")

    def check_content_safety(self, text: str) -> Tuple[bool, str]:
        """
        检查内容是否可能触发内容审核失败

        Args:
            text: 要检查的内容

        Returns:
            (是否安全, 修改后的内容)

        Raises:
            ContentSafetyError: 内容包含敏感词且无法自动修复
        """
        # 早返回：空文本
        if not text:
            return True, text

        # 真正敏感的词汇（只检查明显不当的内容）
        # 注意：不包含"天安门"、"广场"、"故宫"等正常历史文化词汇
        sensitive_keywords: List[str] = [
            # 明显政治敏感（不含正常历史描述）
            "革命",
            "暴动",
            "叛乱",
            "政变",
            # 明显暴力
            "血腥",
            "杀戮",
            "屠杀",
            "武器",
            "枪",
            "刀",
            # 明显色情
            "色情",
            "裸露",
            "情色",
            # 其他明显敏感
            "恐怖",
            "爆炸",
            "毒品",
            "赌博",
        ]

        # 检查是否包含敏感词
        found_keywords: List[str] = [keyword for keyword in sensitive_keywords if keyword in text]

        # 早返回：没有敏感词
        if not found_keywords:
            return True, text

        # 记录发现的敏感词
        Logger.warning(
            "检测到敏感词", logger_name="content_generator", keywords=found_keywords, text_preview=text[:100]
        )

        # 移除敏感词
        modified_text: str = text
        for keyword in found_keywords:
            modified_text = modified_text.replace(keyword, "")

        # 清理多余空格
        modified_text = re.sub(r"\s+", " ", modified_text).strip()

        # 如果修改后的文本太短，抛出异常
        if len(modified_text) < 10:
            raise ContentSafetyError(
                message="内容包含敏感词，移除后内容过短",
                unsafe_content=text[:100],
                matched_keywords=found_keywords,
                details={"original_length": len(text), "modified_length": len(modified_text)},
            )

        return False, modified_text

    def check_and_fix_content_safety(self, content_data: Dict[str, Any], max_retries: int = 3) -> Dict[str, Any]:
        """
        检查并修复内容安全性，如果3次都不行，标记可疑内容

        Args:
            content_data: 生成的内容数据
            max_retries: 最大重试次数

        Returns:
            修复后的内容数据
        """
        for retry_count in range(max_retries):
            has_issue = self._check_and_fix_all_content(content_data, retry_count, max_retries)

            # 早返回：如果没有问题，直接返回
            if not has_issue:
                if retry_count > 0:
                    Logger.info("内容已修复，可以安全使用", logger_name="content_generator")
                return content_data

        # 达到最大重试次数仍有问题，记录可疑内容
        self._save_suspicious_content(content_data)
        return content_data

    def _check_and_fix_all_content(self, content_data: Dict[str, Any], retry_count: int, max_retries: int) -> bool:
        """
        检查并修复所有内容（正文、图片提示词、封面）

        Args:
            content_data: 内容数据
            retry_count: 当前重试次数
            max_retries: 最大重试次数

        Returns:
            是否存在问题
        """
        has_issue = False

        # 检查正文内容
        if self._fix_content_field(content_data, "content", "正文", retry_count, max_retries):
            has_issue = True

        # 检查所有图片提示词
        if self._fix_image_prompts(content_data, retry_count, max_retries):
            has_issue = True

        # 检查封面提示词
        if self._fix_cover_prompt(content_data, retry_count, max_retries):
            has_issue = True

        return has_issue

    def _fix_content_field(
        self, content_data: Dict[str, Any], field: str, field_name: str, retry_count: int, max_retries: int
    ) -> bool:
        """
        修复内容字段

        Args:
            content_data: 内容数据
            field: 字段名
            field_name: 字段显示名称
            retry_count: 当前重试次数
            max_retries: 最大重试次数

        Returns:
            是否存在问题
        """
        content: str = content_data.get(field, "")
        is_safe: bool
        modified_content: str
        is_safe, modified_content = self.check_content_safety(content)

        if is_safe:
            return False

        content_data[field] = modified_content
        Logger.warning(
            f"检测到可疑{field_name}内容，已自动修改",
            logger_name="content_generator",
            retry_attempt=retry_count + 1,
            max_retries=max_retries,
        )
        return True

    def _fix_image_prompts(self, content_data: Dict[str, Any], retry_count: int, max_retries: int) -> bool:
        """
        修复图片提示词

        Args:
            content_data: 内容数据
            retry_count: 当前重试次数
            max_retries: 最大重试次数

        Returns:
            是否存在问题
        """
        has_issue: bool = False
        image_prompts: List[Dict[str, Any]] = content_data.get("image_prompts", [])

        for idx, prompt_data in enumerate(image_prompts):
            prompt = prompt_data.get("prompt", "")
            is_safe, modified_prompt = self.check_content_safety(prompt)

            if not is_safe:
                has_issue = True
                prompt_data["prompt"] = modified_prompt
                Logger.warning(
                    f"检测到可疑提示词（图{idx + 1}），已自动修改",
                    logger_name="content_generator",
                    image_index=idx + 1,
                    retry_attempt=retry_count + 1,
                    max_retries=max_retries,
                )

        return has_issue

    def _fix_cover_prompt(self, content_data: Dict[str, Any], retry_count: int, max_retries: int) -> bool:
        """
        修复封面提示词

        Args:
            content_data: 内容数据
            retry_count: 当前重试次数
            max_retries: 最大重试次数

        Returns:
            是否存在问题
        """
        cover: Dict[str, Any] = content_data.get("cover", {})
        cover_prompt: str = cover.get("prompt", "")

        if not cover_prompt:
            return False

        is_safe: bool
        modified_prompt: str
        is_safe, modified_prompt = self.check_content_safety(cover_prompt)

        if is_safe:
            return False

        cover["prompt"] = modified_prompt
        Logger.warning(
            "检测到可疑封面提示词，已自动修改",
            logger_name="content_generator",
            retry_attempt=retry_count + 1,
            max_retries=max_retries,
        )
        return True

    def _save_suspicious_content(self, content_data: Dict[str, Any]) -> None:
        """
        保存可疑内容到文件

        Args:
            content_data: 内容数据
        """
        Logger.warning(f"经过 {3} 次自动修复，仍有可疑内容", logger_name="content_generator", max_retries=3)

        suspicious_file = os.path.join(self.image_dir, "suspicious_content.txt")

        with open(suspicious_file, "w", encoding="utf - 8") as f:
            f.write("# 可疑内容记录\n\n")
            f.write("以下内容在生成时可能触发内容审核失败，请手动修改后重新生成。\n\n")
            f.write("=" * 60 + "\n\n")

            self._write_suspicious_content(f, content_data.get("content", ""), "正文内容")
            self._write_suspicious_image_prompts(f, content_data.get("image_prompts", []))
            self._write_suspicious_cover(f, content_data.get("cover", {}))

        Logger.info("可疑内容已保存到文件", logger_name="content_generator", file_path="suspicious_content.txt")
        Logger.info("请查看并手动修改后重新运行脚本", logger_name="content_generator")

    def _write_suspicious_content(self, file: Any, content: str, title: str) -> None:
        """写入可疑正文内容"""
        if not content or self.check_content_safety(content)[0]:
            return

        file.write(f"## {title}\n\n")
        file.write(f"```\n{content}\n```\n\n")
        file.write("-" * 60 + "\n\n")

    def _write_suspicious_image_prompts(self, file: Any, image_prompts: List[Dict[str, Any]]) -> None:
        """写入可疑图片提示词"""
        for idx, prompt_data in enumerate(image_prompts):
            prompt = prompt_data.get("prompt", "")
            if prompt and not self.check_content_safety(prompt)[0]:
                file.write(f"## 图{idx + 1}提示词\n\n")
                file.write(f"```\n{prompt}\n```\n\n")
                file.write("-" * 60 + "\n\n")

    def _write_suspicious_cover(self, file: Any, cover: Dict[str, Any]) -> None:
        """写入可疑封面提示词"""
        cover_prompt = cover.get("prompt", "")
        if cover_prompt and not self.check_content_safety(cover_prompt)[0]:
            file.write("## 封面提示词\n\n")
            file.write(f"```\n{cover_prompt}\n```\n\n")
            file.write("-" * 60 + "\n\n")

    def setup_paths(self) -> None:
        """设置路径"""
        # 确保输出目录存在
        output_excel: str = self.config_manager.get("output_excel")
        excel_dir: str = os.path.dirname(output_excel)
        if excel_dir and not os.path.exists(excel_dir):
            os.makedirs(excel_dir)

        # 创建图片输出目录（以日期命名）
        today: str = datetime.now().strftime("%Y%m%d")
        output_image_dir: str = self.config_manager.get("output_image_dir")
        self.image_dir = os.path.join(output_image_dir, today)
        if not os.path.exists(self.image_dir):
            os.makedirs(self.image_dir)
            Logger.info("已创建图片目录", logger_name="content_generator", directory=self.image_dir)

    def read_input_file(self) -> str:
        """
        读取输入文档

        Returns:
            文件内容

        Raises:
            CustomFileNotFoundError: 文件不存在
            ContentValidationError: 文件内容为空
        """
        input_path = self.config_manager.get("input_file")

        if not os.path.exists(input_path):
            raise CustomFileNotFoundError(
                file_path=input_path, suggestion="请确保输入文件存在，或在配置文件中指定正确的路径"
            )

        try:
            with open(input_path, "r", encoding="utf - 8") as f:
                content = f.read().strip()
        except Exception as e:
            raise wrap_exception(
                e,
                message=f"读取输入文件失败: {input_path}",
                exception_class=CustomFileNotFoundError,
                file_path=input_path,
                operation="read",
            )

        if not content:
            raise ContentValidationError(
                message=f"输入文件为空: {input_path}",
                field="input_file",
                validation_rule="non_empty",
                details={"file_path": input_path},
            )

        Logger.info(
            "已读取输入文件", logger_name="content_generator", file_path=input_path, content_length=len(content)
        )
        return content

    def generate_content(self, raw_content: str) -> Dict[str, Any]:
        """
        调用AI生成小红书文案和绘画提示词，包含 3 次重写逻辑。
        支持缓存功能，避免重复生成相同内容。
        支持速率限制，避免超过 API 配额。

        Args:
            raw_content: 原始输入内容

        Returns:
            生成的内容数据字典
        """
        # 1. 检查缓存
        cached_result = self._check_cache(raw_content)
        if cached_result is not None:
            return cached_result

        # 2. 初始化 OpenAI 客户端
        client, model = self._initialize_openai_client()

        # 3. 迭代生成内容（包含自我评估）
        best_result = self._generate_with_iterations(client, model, raw_content)

        # 4. 安全检查
        Logger.info("正在检查生成内容的安全性", logger_name="content_generator")
        best_result = self.check_and_fix_content_safety(best_result)

        # 5. 保存到缓存
        self._save_to_cache(raw_content, best_result)

        Logger.info("AI内容生成成功", logger_name="content_generator")
        return best_result

    def _call_openai_with_rate_limit(
        self, 
        client: openai.OpenAI, 
        model: str, 
        messages: List[Dict[str, str]], 
        temperature: float = 0.8, 
        response_format: Optional[Dict[str, str]] = None
    ) -> Any:
        """
        调用 OpenAI API 并应用速率限制

        Args:
            client: OpenAI 客户端实例
            model: 模型名称
            messages: 消息列表
            temperature: 温度参数
            response_format: 响应格式

        Returns:
            API 响应对象
        """
        return self.api_handler.call_openai(
            client=client, model=model, messages=messages, temperature=temperature, response_format=response_format
        )

    def _build_generation_prompt(self, raw_content: str, attempt: int = 1) -> str:
        """构建生成提示词"""
        return f"""# Role: 老北京文化·小红书金牌运营 & 视觉导演

    ## Goals
    1. 读取用户输入的原始内容。
    2. 改写为具备"爆款潜质"的小红书文案。文案必须充满生活气息，避免总结性、AI感的陈述，多用细节描写。
    3. 生成 3 - 5 组 AI 绘画提示词。

    ## Constraints
    - **文字风格**：必须地道，多用短句，多用Emoji。拒绝"总分总"的枯燥结构。
    - **画面风格**：90年代北京纪实，胶片质感。
    - **牌匾文字**：如果涉及故宫牌匾，请明确要求文字为"建极绥猷"，并描述其颜色（蓝底金字）。

    ## Workflow
    ### Step 1: 文案创作
    - 请提供 5 个【标题】。
    - 正文：开头要抓人，中间要动人，结尾要有互动。

    ### Step 2: 画面提取
    - 包含至少 4 张故事图提示词。
    - 牌匾策略：针对包含牌匾的图，在 Prompt 中强制加入"建极绥猷 (Jian Ji Sui You)"字样。

    ## Output Format
    {{{{
      "titles": ["...", "..."],
      "content": "...",
      "tags": "...",
      "image_prompts": [
        {{{{"scene": "...", "prompt": "..."}}}},
        ...
      ],
      "cover": {{{{"scene": "...", "title": "...", "prompt": "..."}}}}
    }}}}

    ## 原始内容：
    {raw_content}
    """
    def _check_cache(self, raw_content: str) -> Optional[Dict[str, Any]]:
        """
        检查缓存中是否存在结果

        Args:
            raw_content: 原始输入内容

        Returns:
            缓存的结果，如果不存在则返回 None
        """
        # 早返回：缓存未启用
        if not self._cache_enabled or self.cache is None:
            return None

        cache_key = self._generate_cache_key(raw_content)
        cached_result = self.cache.get(cache_key)

        # 早返回：缓存未命中
        if cached_result is None:
            Logger.info(
                "缓存未命中，开始生成新内容", logger_name="content_generator", cache_key=cache_key[:16] + "..."
            )
            return None

        # 缓存命中
        Logger.info(
            "✅ 缓存命中，直接返回缓存结果", logger_name="content_generator", cache_key=cache_key[:16] + "..."
        )

        cache_stats = self.cache.get_stats()
        Logger.debug("缓存统计", logger_name="content_generator", **cache_stats)
        return cached_result

    def _save_to_cache(self, raw_content: str, result: Dict[str, Any]) -> None:
        """
        保存结果到缓存

        Args:
            raw_content: 原始输入内容
            result: 生成的结果
        """
        # 早返回：缓存未启用
        if not self._cache_enabled or self.cache is None:
            return

        cache_key = self._generate_cache_key(raw_content)
        self.cache.set(cache_key, result)

        Logger.info("✅ 生成结果已保存到缓存", logger_name="content_generator", cache_key=cache_key[:16] + "...")

        cache_stats = self.cache.get_stats()
        Logger.debug("缓存统计", logger_name="content_generator", **cache_stats)

    def _initialize_openai_client(self) -> Tuple[openai.OpenAI, str]:
        """
        初始化 OpenAI 客户端

        Returns:
            (客户端实例, 模型名称)
        """
        api_key = self.config_manager.get("openai_api_key") or os.getenv("OPENAI_API_KEY")

        # 早返回：API Key 不存在
        if not api_key:
            raise ValueError("❌ 未找到 API Key")

        base_url = self.config_manager.get("openai_base_url")
        model = self.config_manager.get("openai_model", "gpt - 4")

        # 处理 Qwen 模型兼容性
        base_url, model = self._handle_qwen_compatibility(base_url, model)

        # 构建客户端参数
        client_kwargs = {"api_key": api_key}
        if base_url:
            client_kwargs["base_url"] = base_url

        return openai.OpenAI(**client_kwargs), model

    def _handle_qwen_compatibility(self, base_url: Optional[str], model: str) -> Tuple[Optional[str], str]:
        """
        处理 Qwen 模型的兼容性配置

        Args:
            base_url: 原始 base_url
            model: 原始模型名称

        Returns:
            (处理后的 base_url, 处理后的模型名称)
        """
        # 早返回：不是 Qwen 模型
        if not (model == "qwen" or (isinstance(model, str) and model.startswith("qwen-"))):
            return base_url, model

        # 设置默认 base_url
        if not base_url:
            base_url = "https://dashscope.aliyuncs.com/compatible-mode/v1"

        # 转换模型名称
        if model == "qwen":
            model = "qwen-plus"

        return base_url, model

    def _generate_initial_content(self, client: openai.OpenAI, model: str, raw_content: str) -> Dict[str, Any]:
        """
        生成初始内容

        Args:
            client: OpenAI 客户端
            model: 模型名称
            raw_content: 原始输入内容

        Returns:
            生成的内容字典
        """
        response = self._call_openai_with_rate_limit(
            client=client,
            model=model,
            messages=[
                {"role": "system", "content": "你是一位专业的小红书内容创作专家。请严格按照JSON格式输出。"},
                {"role": "user", "content": self._build_generation_prompt(raw_content)},
            ],
            temperature=0.8,
            response_format={"type": "json_object"},
        )

        result_text = response.choices[0].message.content.strip()
        return json.loads(result_text)

    def _evaluate_content(self, client: openai.OpenAI, model: str, content: str) -> str:
        """
        评估生成的内容质量

        Args:
            client: OpenAI 客户端
            model: 模型名称
            content: 待评估的内容

        Returns:
            评估反馈
        """
        eval_prompt = f"""请作为资深主编审阅以下小红书文案：
    ---
    {content}
    ---
    评价该文案是否符合：
    1. 京味儿是否地道？
    2. 情感是否细腻？
    3. 排版是否舒适？
    4. 是否通过"叙事"而不是"说教"？

    如果评价为"优秀"，请直接返回"PASS"。
    如果需要优化，请指出不足，并给出修改意见。"""

        eval_response = self._call_openai_with_rate_limit(
            client=client,
            model=model,
            messages=[
                {"role": "system", "content": "你是一位极其挑剔的小红书内容主编。"},
                {"role": "user", "content": eval_prompt},
            ],
            temperature=0.5,
        )

        return eval_response.choices[0].message.content.strip()

    def _generate_with_iterations(
        self, client: openai.OpenAI, model: str, raw_content: str, max_attempts: int = 3
    ) -> Dict[str, Any]:
        """
        迭代生成内容，包含自我评估和改进

        Args:
            client: OpenAI 客户端
            model: 模型名称
            raw_content: 原始输入内容
            max_attempts: 最大尝试次数

        Returns:
            最佳生成结果
        """

        def evaluator(result: Dict[str, Any]) -> Tuple[bool, str]:
            """评估生成结果"""
            eval_feedback = self._evaluate_content(client, model, result.get("content", ""))

            # 早返回：评估通过
            if "PASS" in eval_feedback.upper():
                return False, ""  # 不需要继续

            # 评估未通过，记录反馈
            Logger.info(
                f"主编反馈：{eval_feedback[:100]}...",
                logger_name="content_generator",
                feedback_preview=eval_feedback[:100],
            )
            return True, eval_feedback  # 需要继续

        return self.api_handler.call_openai_with_evaluation(
            client=client,
            model=model,
            raw_content=raw_content,
            prompt_builder=self._build_generation_prompt,
            max_iterations=max_attempts,
            evaluator=evaluator,
        )

    def save_to_excel(self, content_data: Dict[str, Any], raw_content: str) -> None:
        """
        保存内容到Excel文件

        Args:
            content_data: 生成的内容数据
            raw_content: 原始输入内容
        """
        excel_path: str = self.config_manager.get("output_excel")
        headers: List[str] = [
            "生成时间",
            "原始内容",
            "标题1",
            "标题2",
            "标题3",
            "标题4",
            "标题5",
            "正文内容",
            "标签",
            "图片提示词1",
            "图片提示词2",
            "图片提示词3",
            "图片提示词4",
            "封面标题",
            "封面提示词",
            "图片保存路径",
        ]

        # 检查文件是否存在
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "小红书内容"

            # 创建表头

            # 设置表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 设置列宽
            column_widths = [18, 40, 30, 30, 30, 30, 30, 60, 40, 50, 50, 50, 50, 30, 50, 30]
            for col_idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 添加新行
        now: str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row_data: List[Any] = [
            now,  # 生成时间
            raw_content[:500] if len(raw_content) > 500 else raw_content,  # 原始内容（截断）
        ]

        # 添加标题
        titles: List[str] = content_data.get("titles", [])
        for i in range(5):
            row_data.append(titles[i] if i < len(titles) else "")

        # 添加正文和标签
        row_data.append(content_data.get("content", ""))
        row_data.append(content_data.get("tags", ""))

        # 添加图片提示词（至少4张故事图）
        image_prompts: List[Dict[str, Any]] = content_data.get("image_prompts", [])
        for i in range(4):
            if i < len(image_prompts):
                prompt_text: str = f"{image_prompts[i].get('scene', '')}: {image_prompts[i].get('prompt', '')}"
                row_data.append(prompt_text)
            else:
                row_data.append("")

        # 封面标题、封面提示词
        cover: Dict[str, Any] = content_data.get("cover", {})
        row_data.append(cover.get("title", ""))
        row_data.append(cover.get("prompt", ""))

        # 添加图片保存路径
        row_data.append(self.image_dir)

        # 写入数据
        ws.append(row_data)

        # 设置数据行样式
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 保存文件
        wb.save(excel_path)
        Logger.info("内容已保存到Excel", logger_name="content_generator", file_path=excel_path)

    def save_image_prompts(self, content_data: Dict[str, Any]) -> None:
        """
        保存图片提示词到文件：4 张故事图 + 1 张封面（带短标题）
        """
        prompts_file: str = os.path.join(self.image_dir, "image_prompts.txt")

        with open(prompts_file, "w", encoding="utf - 8") as f:
            f.write("# AI绘画提示词\n\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            # 保存正文内容（用于后续分段叠加到图片上）
            content: str = content_data.get("content", "").strip()
            if content:
                f.write("## 正文内容\n\n")
                f.write(f"{content}\n\n")
                f.write("---\n\n")

            # 故事图：至少 4 张
            image_prompts: List[Dict[str, Any]] = content_data.get("image_prompts", [])[:4]
            for idx, prompt_data in enumerate(image_prompts, start=1):
                f.write(f"## 图{idx}: {prompt_data.get('scene', '')}\n\n")
                f.write(f"```\n{prompt_data.get('prompt', '')}\n```\n\n")

            # 封面：短标题 + 带标题的 prompt
            cover: Dict[str, Any] = content_data.get("cover", {})
            if cover.get("title") and cover.get("prompt"):
                f.write(f"## 封面: {cover.get('title', '')}\n\n")
                f.write(f"```\n{cover.get('prompt', '')}\n```\n\n")

        Logger.info("图片提示词已保存", logger_name="content_generator", file_path=prompts_file)

    def save_full_content(self, content_data: Dict[str, Any], raw_content: str) -> None:
        """
        保存完整内容到Markdown文件

        Args:
            content_data: 生成的内容数据
            raw_content: 原始输入内容
        """
        md_file: str = os.path.join(self.image_dir, "content.md")

        with open(md_file, "w", encoding="utf - 8") as f:
            f.write("# 小红书文案预览\n\n")
            f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            f.write("## 📕 可选标题\n\n")
            titles: List[str] = content_data.get("titles", [])
            for idx, title in enumerate(titles, start=1):
                f.write(f"{idx}. {title}\n")

            f.write("\n## 📝 正文内容\n\n")
            f.write(content_data.get("content", ""))

            f.write("\n\n## 🏷️ 标签\n\n")
            f.write(content_data.get("tags", ""))

            f.write("\n\n## 🎨 AI绘画提示词\n\n")
            image_prompts: List[Dict[str, Any]] = content_data.get("image_prompts", [])[:4]
            for idx, prompt_data in enumerate(image_prompts, start=1):
                f.write(f"### 图{idx}: {prompt_data.get('scene', '')}\n\n")
                f.write(f"```\n{prompt_data.get('prompt', '')}\n```\n\n")
            cover: Dict[str, Any] = content_data.get("cover", {})
            if cover.get("title") and cover.get("prompt"):
                f.write(f"### 封面: {cover.get('title', '')}\n\n")
                f.write(f"```\n{cover.get('prompt', '')}\n```\n\n")

            f.write("\n---\n\n")
            f.write("## 📄 原始输入内容\n\n")
            f.write(raw_content)

        Logger.info("完整内容已保存", logger_name="content_generator", file_path=md_file)

    def generate_single_content(self, input_text: str) -> Dict[str, Any]:
        """
        为Web API生成单条内容

        Args:
            input_text: 输入文本

        Returns:
            包含title, content, tags, image_prompt的字典
        """
        try:
            # 调用generate_content生成完整内容
            content_data: Dict[str, Any] = self.generate_content(input_text)

            # 提取第一个标题
            titles: List[str] = content_data.get("titles", [])
            title: str = titles[0] if titles else "老北京记忆"

            # 提取正文
            content: str = content_data.get("content", "")

            # 提取标签（转换为列表）
            tags_str: str = content_data.get("tags", "")
            tags: List[str] = [tag.strip().replace("#", "") for tag in tags_str.split("#") if tag.strip()]

            # 提取第一个图片提示词
            image_prompts: List[Dict[str, Any]] = content_data.get("image_prompts", [])
            image_prompt: str = ""
            if image_prompts:
                first_prompt: Dict[str, Any] = image_prompts[0]
                image_prompt = first_prompt.get("prompt", "")

            return {
                "title": title,
                "content": content,
                "tags": tags,
                "image_prompt": image_prompt,
                "raw_data": content_data,  # 保留原始数据以便需要时使用
            }

        except Exception as e:
            ErrorHandler.handle_error(
                error=e,
                logger_name="content_generator",
                operation_name="单条内容生成",
                context={"input_length": len(input_text)},
            )

    def run(self) -> None:
        """运行主流程"""
        try:
            Logger.info("=" * 60, logger_name="content_generator")
            Logger.info("老北京文化·小红书内容生成器", logger_name="content_generator")
            Logger.info("=" * 60, logger_name="content_generator")

            # 1. 读取输入文件
            raw_content = self.read_input_file()

            # 2. 生成内容
            Logger.info("正在调用AI生成内容", logger_name="content_generator")
            content_data = self.generate_content(raw_content)

            # 3. 保存到Excel
            Logger.info("正在保存到Excel", logger_name="content_generator")
            self.save_to_excel(content_data, raw_content)

            # 4. 保存图片提示词
            Logger.info("正在保存图片提示词", logger_name="content_generator")
            self.save_image_prompts(content_data)

            # 5. 保存完整内容
            Logger.info("正在保存完整内容", logger_name="content_generator")
            self.save_full_content(content_data, raw_content)

            Logger.info("=" * 60, logger_name="content_generator")
            Logger.info("所有任务完成！", logger_name="content_generator")
            Logger.info(f"Excel文件: {self.config_manager.get('output_excel')}", logger_name="content_generator")
            Logger.info(f"图片目录: {self.image_dir}", logger_name="content_generator")
            Logger.info("=" * 60, logger_name="content_generator")

        except Exception as e:
            ErrorHandler.handle_error(error=e, logger_name="content_generator", operation_name="主流程运行")


def main() -> None:
    """主函数"""
    import argparse
    from src.core.config_manager import ConfigManager

    parser = argparse.ArgumentParser(description="老北京文化·小红书内容生成器")
    parser.add_argument("-c", "--config", default="config/config.json", help="配置文件路径 (默认: config/config.json)")

    args = parser.parse_args()

    # 使用 ConfigManager 加载配置
    config_manager = ConfigManager(args.config)
    generator = RedBookContentGenerator(config_manager=config_manager)
    generator.run()


if __name__ == "__main__":
    main()
