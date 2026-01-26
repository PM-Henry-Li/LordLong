#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
老北京文化·小红书内容生成器
读取文档内容，生成小红书文案和AI绘画提示词，保存到Excel和文件夹
"""

import os
import json
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import openai
from typing import List, Dict, Tuple
import re


class RedBookContentGenerator:
    """小红书内容生成器"""
    
    def __init__(self, config_path: str = "config.json"):
        """
        初始化生成器
        
        Args:
            config_path: 配置文件路径
        """
        self.config = self._load_config(config_path)
        self.setup_paths()
        
        # API Key会在调用时检查，这里不需要初始化
    
    def _load_config(self, config_path: str) -> dict:
        """加载配置文件"""
        default_config = {
            "input_file": "input_content.txt",
            "output_excel": "output/redbook_content.xlsx",
            "output_image_dir": "output/images",
            "openai_api_key": "",
            "openai_model": "gpt-4",
            "openai_base_url": None
        }
        
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                user_config = json.load(f)
                default_config.update(user_config)
        else:
            # 创建默认配置文件
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, ensure_ascii=False, indent=2)
            print(f"✅ 已创建默认配置文件: {config_path}")
        
        return default_config
    
    def check_content_safety(self, text: str) -> Tuple[bool, str]:
        """
        检查内容是否可能触发内容审核失败
        
        Args:
            text: 要检查的内容
            
        Returns:
            (是否安全, 修改后的内容)
        """
        if not text:
            return True, text
        
        import re
        
        # 真正敏感的词汇（只检查明显不当的内容）
        # 注意：不包含"天安门"、"广场"、"故宫"等正常历史文化词汇
        sensitive_keywords = [
            # 明显政治敏感（不含正常历史描述）
            '革命', '暴动', '叛乱', '政变',
            # 明显暴力
            '血腥', '杀戮', '屠杀', '武器', '枪', '刀',
            # 明显色情
            '色情', '裸露', '情色',
            # 其他明显敏感
            '恐怖', '爆炸', '毒品', '赌博',
        ]
        
        # 检查是否包含敏感词
        # 注意：只检查明显敏感的词，不误杀正常历史文化内容
        found_keywords = []
        for keyword in sensitive_keywords:
            if keyword in text:
                found_keywords.append(keyword)
        
        if found_keywords:
            # 尝试移除敏感词
            modified_text = text
            for keyword in found_keywords:
                modified_text = modified_text.replace(keyword, '')
            # 清理多余空格
            modified_text = re.sub(r'\s+', ' ', modified_text).strip()
            return False, modified_text
        
        return True, text
    
    def check_and_fix_content_safety(self, content_data: Dict, max_retries: int = 3) -> Dict:
        """
        检查并修复内容安全性，如果3次都不行，标记可疑内容
        
        Args:
            content_data: 生成的内容数据
            max_retries: 最大重试次数
            
        Returns:
            修复后的内容数据
        """
        retry_count = 0
        suspicious_items = []
        
        while retry_count < max_retries:
            has_issue = False
            
            # 检查正文内容
            content = content_data.get("content", "")
            is_safe, modified_content = self.check_content_safety(content)
            if not is_safe:
                has_issue = True
                content_data["content"] = modified_content
                print(f"  ⚠️  检测到可疑正文内容，已自动修改（尝试 {retry_count + 1}/{max_retries}）")
            
            # 检查所有图片提示词
            image_prompts = content_data.get("image_prompts", [])
            for idx, prompt_data in enumerate(image_prompts):
                prompt = prompt_data.get("prompt", "")
                is_safe, modified_prompt = self.check_content_safety(prompt)
                if not is_safe:
                    has_issue = True
                    prompt_data["prompt"] = modified_prompt
                    print(f"  ⚠️  检测到可疑提示词（图{idx+1}），已自动修改（尝试 {retry_count + 1}/{max_retries}）")
            
            # 检查封面提示词
            cover = content_data.get("cover", {})
            cover_prompt = cover.get("prompt", "")
            if cover_prompt:
                is_safe, modified_prompt = self.check_content_safety(cover_prompt)
                if not is_safe:
                    has_issue = True
                    cover["prompt"] = modified_prompt
                    print(f"  ⚠️  检测到可疑封面提示词，已自动修改（尝试 {retry_count + 1}/{max_retries}）")
            
            # 如果没有问题，返回
            if not has_issue:
                if retry_count > 0:
                    print(f"  ✅ 内容已修复，可以安全使用")
                return content_data
            
            retry_count += 1
            
            # 如果还有问题且已达到最大重试次数，记录可疑内容
            if retry_count >= max_retries and has_issue:
                print(f"  ⚠️  经过 {max_retries} 次自动修复，仍有可疑内容")
                # 记录可疑内容
                suspicious_file = os.path.join(self.image_dir, "suspicious_content.txt")
                with open(suspicious_file, 'w', encoding='utf-8') as f:
                    f.write("# 可疑内容记录\n\n")
                    f.write("以下内容在生成时可能触发内容审核失败，请手动修改后重新生成。\n\n")
                    f.write("=" * 60 + "\n\n")
                    
                    if content and not self.check_content_safety(content)[0]:
                        f.write("## 正文内容\n\n")
                        f.write(f"```\n{content}\n```\n\n")
                        f.write("-" * 60 + "\n\n")
                    
                    for idx, prompt_data in enumerate(image_prompts):
                        prompt = prompt_data.get("prompt", "")
                        if prompt and not self.check_content_safety(prompt)[0]:
                            f.write(f"## 图{idx+1}提示词\n\n")
                            f.write(f"```\n{prompt}\n```\n\n")
                            f.write("-" * 60 + "\n\n")
                    
                    cover_prompt = cover.get("prompt", "")
                    if cover_prompt and not self.check_content_safety(cover_prompt)[0]:
                        f.write("## 封面提示词\n\n")
                        f.write(f"```\n{cover_prompt}\n```\n\n")
                        f.write("-" * 60 + "\n\n")
                
                print(f"  📝 可疑内容已保存到: suspicious_content.txt")
                print(f"  💡 请查看并手动修改后重新运行脚本")
        
        return content_data
    
    def setup_paths(self):
        """设置路径"""
        # 确保输出目录存在
        excel_dir = os.path.dirname(self.config["output_excel"])
        if excel_dir and not os.path.exists(excel_dir):
            os.makedirs(excel_dir)
        
        # 创建图片输出目录（以日期命名）
        today = datetime.now().strftime("%Y%m%d")
        self.image_dir = os.path.join(self.config["output_image_dir"], today)
        if not os.path.exists(self.image_dir):
            os.makedirs(self.image_dir)
            print(f"✅ 已创建图片目录: {self.image_dir}")
    
    def read_input_file(self) -> str:
        """读取输入文档"""
        input_path = self.config["input_file"]
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"❌ 输入文件不存在: {input_path}")
        
        with open(input_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()
        
        if not content:
            raise ValueError(f"❌ 输入文件为空: {input_path}")
        
        print(f"✅ 已读取输入文件: {input_path} ({len(content)} 字符)")
        return content
    
    def generate_content(self, raw_content: str) -> Dict:
        """
        调用AI生成小红书文案和绘画提示词
        
        Args:
            raw_content: 原始文档内容
            
        Returns:
            包含标题、正文、标签、绘画提示词的字典
        """
        prompt = f"""# Role: 老北京文化·小红书金牌运营 & 视觉导演

## Profile
你是一位深耕"老北京记忆"领域的小红书博主，擅长用细腻、怀旧、有温度的笔触重现四九城的往事。同时，你也是一位AI绘画提示词专家，能够将文字画面转化为风格统一的视觉描述。

## Goals
1. 读取用户输入的原始文案（通常是片段式的老北京回忆）。
2. 将其改写为一篇具备"爆款潜质"的小红书文案。
3. 提取文案中的关键画面，生成 3-5 组风格高度统一的 AI 绘画提示词（用于 Nano Banana 或 Stable Diffusion）。

## Constraints & Style
1. **文案风格**：
   - **京味儿**：适当使用北京方言（如：这地界儿、发小儿、甚至、大概齐），但不要过重影响阅读。
   - **沉浸感**：强调感官描写（鸽哨声、冬储大白菜味、煤球味、槐花香）。
   - **情感共鸣**：引发"回不去的小时候"或"岁月静好"的共鸣。
   - **排版**：多分段，每段不超过3行，多用Emoji，视觉舒适。

2. **视觉风格（必须统一）**：
   - **核心基调**：**90年代北京纪实摄影 (90s Beijing Documentary Photography)**。
   - **胶片质感**：模拟 **Kodak Vision3 500T** 或 **Fujifilm Superia** 胶卷色彩，带有细腻的颗粒感 (Subtle Film Grain) 和 宽容度高的光影。
   - **光影氛围**：偏好 **暖色调 (Warm Tone)**，如夕阳余晖 (Golden Hour)、老式白炽灯光、冬日暖阳。避免过于冷冽或现代的高对比度霓虹感。
   - **画面细节**：
     - 环境：老旧的红砖墙、斑驳的木门、甚至胡同里的杂物（如堆放的大白菜、停靠的二八大杠自行车）。
     - 人物：穿着90年代特色的服装（如军大衣、毛衣、运动校服），表情自然朴实。
     - 构图：采用 **中焦段 (35mm-50mm)**，既交代环境由于有主体，避免过于广角畸变。

## Workflow

### Step 1: 文案创作
请提供 5 个吸引人的**【标题】**（包含悬念、情感或特定地名）。
正文请按以下结构撰写：
- **开头**：用一个具体的场景或声音切入，瞬间拉回那个年代。
- **中间**：展开故事，加入感官细节。请使用**连贯的叙事风格**，而不是碎片的句子。
- **结尾**：升华情感，引导互动（问问大家还记不记得）。
- **标签**：添加 #老北京 #胡同记忆 #胶片 #童年回忆 等相关Tag。
- **重要格式要求**：文中需要换行的地方，请直接使用**标准的换行符 (\\n)**，**严禁**使用转义的 `\\\\n` 或 `\\\\\\\\n`。确保输出的 JSON 字符串可以直接被 Python 解析出正确的换行。

### Step 2: 画面提取 (AI Image Prompts)
- **故事图**：基于改写后的文案，提取 **至少 4 个**最具画面感的场景（必须 ≥4 个）。
- **【特殊要求：牌匾文字精准还原】**：如果画面中涉及“太和殿”、“牌匾”等场景，请务必在 Prompt 中明确指定牌匾上的四个大字为 **"建极绥猷"** (Traditional Chinese: 建極綏猷)。描述其为“金色木制牌匾，蓝色底色，遒劲有力的皇家楷书书法”。
- **重要提示**：故事图的正文内容分段会叠加到图片底部，**每段文字建议控制在50-80字以内**，确保能在3行内完整显示，避免文字被截断或重叠。
- **封面图**：额外生成 1 张适合小红书的**封面图**，要求：
  - 画面符合主题故事、适合做笔记封面；
  - 由你根据主题创作一句吸引人的短标题，存储在cover.title中。**标题长度建议控制在8-10字以内**，避免过长导致显示时超出图片范围。如果主题需要较长标题，可以适当精简或使用更简洁的表达。
  - **重要说明**：封面图的画面**不需要包含文字**，只需要生成适合做封面的背景画面。文字标题会通过后期处理叠加到图片上，确保文字100%准确。因此，cover.prompt中**不要描述文字内容**，只描述画面构图、氛围和风格即可。
  - 画面要求：适合小红书封面，构图美观，在画面顶部或中央区域留出空间（用于后续叠加文字），色调与风格与故事图保持一致。

输出格式为**中文 Prompt**，但为了更好的生成效果，请在中文描述后附带关键的**英文风格词**。
必须包含以下**固定风格关键词**：
*Fixed Style Keywords: 90s Beijing street photography, vintage Kodak film look, nostalgic warm tone, cinematic lighting, photorealistic, highly detailed, 8k resolution, 3:4 aspect ratio*

## Output Format
请严格按照以下JSON格式输出，不要包含任何其他文字：

{{
  "titles": ["标题1", "标题2", "标题3", "标题4", "标题5"],
  "content": "正文内容（带Emoji，多分段）",
  "tags": "#老北京 #胡同记忆 #胶片 #童年回忆 #...",
  "image_prompts": [
    {{ "scene": "场景简述", "prompt": "完整的中文Prompt，包含风格关键词" }},
    {{ "scene": "场景简述", "prompt": "完整的中文Prompt，包含风格关键词" }},
    {{ "scene": "场景简述", "prompt": "完整的中文Prompt，包含风格关键词" }},
    {{ "scene": "场景简述", "prompt": "完整的中文Prompt，包含风格关键词" }}
  ],
  "cover": {{
    "scene": "封面画面简述（适合小红书封面的构图与氛围）",
    "title": "短标题（中文，6–12字，将醒目显示在封面图上）",
    "prompt": "中文Prompt。包含详细的画面描述、构图描述、氛围描述 and 固定风格关键词。要求确保画面适合叠加文字。"
  }}
}}

注意：image_prompts 至少 4 条；所有 prompt 均使用中文描述；cover.prompt **不需要描述文字内容**（文字会通过后期处理叠加）。

## 用户输入的原始文案：
{raw_content}

请开始生成内容："""

    def generate_content(self, raw_content: str) -> Dict:
        """
        调用AI生成小红书文案和绘画提示词，包含 3 次重写逻辑。
        """
        # 获取基础配置
        api_key = self.config.get("openai_api_key") or os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("❌ 未找到 API Key")
        
        base_url = self.config.get("openai_base_url")
        model = self.config.get("openai_model", "gpt-4")
        
        # 兼容性处理
        if model == "qwen" or (isinstance(model, str) and model.startswith("qwen-")):
            if not base_url:
                base_url = "https://dashscope.aliyuncs.com/compatible-mode/v1"
            if model == "qwen":
                model = "qwen-plus"
        
        client_kwargs = {"api_key": api_key}
        if base_url:
            client_kwargs["base_url"] = base_url
        client = openai.OpenAI(**client_kwargs)

        best_result = None
        max_rewrite_attempts = 3
        
        for attempt in range(1, max_rewrite_attempts + 1):
            print(f"\n🤖 正在尝试生成内容 (第 {attempt}/{max_rewrite_attempts} 次)...")
            
            try:
                # 1. 初步生成
                response = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "你是一位专业的小红书内容创作专家。请严格按照JSON格式输出。"},
                        {"role": "user", "content": self._build_generation_prompt(raw_content)}
                    ],
                    temperature=0.8,
                    response_format={"type": "json_object"}
                )
                result_text = response.choices[0].message.content.strip()
                result = json.loads(result_text)
                
                # 2. 自我评估与改写逻辑
                if attempt < max_rewrite_attempts:
                    eval_prompt = f"""请作为资深主编审阅以下小红书文案：
---
{result.get('content', '')}
---
评价该文案是否符合：
1. 京味儿是否地道？
2. 情感是否细腻？
3. 排版是否舒适？
4. 是否通过“叙事”而不是“说教”？

如果评价为“优秀”，请直接返回“PASS”。
如果需要优化，请指出不足，并给出修改意见。"""
                    
                    eval_response = client.chat.completions.create(
                        model=model,
                        messages=[
                            {"role": "system", "content": "你是一位极其挑剔的小红书内容主编。"},
                            {"role": "user", "content": eval_prompt}
                        ],
                        temperature=0.5
                    )
                    eval_feedback = eval_response.choices[0].message.content.strip()
                    
                    if "PASS" in eval_feedback.upper():
                        print(f"  ✨ 文案质量优秀，通过审核。")
                        best_result = result
                        break
                    else:
                        print(f"  📝 主编反馈：{eval_feedback[:100]}...")
                        # 准备下一次生成的 prompt
                        raw_content = f"{raw_content}\n\n[上一次生成的不足之处及改进意见]：{eval_feedback}"
                        best_result = result # 先存一个保底
                else:
                    best_result = result
                    
            except Exception as e:
                print(f"  ❌ 第 {attempt} 次生成失败: {e}")
                if attempt == max_rewrite_attempts and not best_result:
                    raise
        
        # 3. 结果验证与安全检查
        if not best_result:
            raise ValueError("❌ 无法生成有效内容")
            
        print("\n🔍 正在检查生成内容的安全性...")
        best_result = self.check_and_fix_content_safety(best_result)
        print("✅ AI内容生成成功")
        return best_result

    def _build_generation_prompt(self, raw_content: str) -> str:
        """构建生成提示词"""
        return f"""# Role: 老北京文化·小红书金牌运营 & 视觉导演

## Goals
1. 读取用户输入的原始内容。
2. 改写为具备"爆款潜质"的小红书文案。文案必须充满生活气息，避免总结性、AI感的陈述，多用细节描写。
3. 生成 3-5 组 AI 绘画提示词。

## Constraints
- **文字风格**：必须地道，多用短句，多用Emoji。拒绝“总分总”的枯燥结构。
- **画面风格**：90年代北京纪实，胶片质感。
- **牌匾文字**：如果涉及故宫牌匾，请明确要求文字为“建极绥猷”，并描述其颜色（蓝底金字）。

## Workflow
### Step 1: 文案创作
- 请提供 5 个【标题】。
- 正文：开头要抓人，中间要动人，结尾要有互动。

### Step 2: 画面提取
- 包含至少 4 张故事图提示词。
- 牌匾策略：针对包含牌匾的图，在 Prompt 中强制加入“建极绥猷 (Jian Ji Sui You)”字样。

## Output Format
{{
  "titles": ["...", "..."],
  "content": "...",
  "tags": "...",
  "image_prompts": [
    {{ "scene": "...", "prompt": "..." }},
    ...
  ],
  "cover": {{ "scene": "...", "title": "...", "prompt": "..." }}
}}

## 原始内容：
{raw_content}
"""
    
    def save_to_excel(self, content_data: Dict, raw_content: str):
        """
        保存内容到Excel文件
        
        Args:
            content_data: 生成的内容数据
            raw_content: 原始输入内容
        """
        excel_path = self.config["output_excel"]
        headers = [
            "生成时间", "原始内容", "标题1", "标题2", "标题3", "标题4", "标题5",
            "正文内容", "标签", "图片提示词1", "图片提示词2", "图片提示词3", "图片提示词4",
            "封面标题", "封面提示词", "图片保存路径"
        ]
        
        # 检查文件是否存在
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "小红书内容"
            
            # 创建表头
            
            # 设置表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置列宽
            column_widths = [18, 40, 30, 30, 30, 30, 30, 60, 40, 50, 50, 50, 50, 30, 50, 30]
            for col_idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 添加新行
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row_data = [
            now,  # 生成时间
            raw_content[:500] if len(raw_content) > 500 else raw_content,  # 原始内容（截断）
        ]
        
        # 添加标题
        titles = content_data.get("titles", [])
        for i in range(5):
            row_data.append(titles[i] if i < len(titles) else "")
        
        # 添加正文和标签
        row_data.append(content_data.get("content", ""))
        row_data.append(content_data.get("tags", ""))
        
        # 添加图片提示词（至少4张故事图）
        image_prompts = content_data.get("image_prompts", [])
        for i in range(4):
            if i < len(image_prompts):
                prompt_text = f"{image_prompts[i].get('scene', '')}: {image_prompts[i].get('prompt', '')}"
                row_data.append(prompt_text)
            else:
                row_data.append("")
        
        # 封面标题、封面提示词
        cover = content_data.get("cover", {})
        row_data.append(cover.get("title", ""))
        row_data.append(cover.get("prompt", ""))
        
        # 添加图片保存路径
        row_data.append(self.image_dir)
        
        # 写入数据
        ws.append(row_data)
        
        # 设置数据行样式
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # 保存文件
        wb.save(excel_path)
        print(f"✅ 内容已保存到Excel: {excel_path}")
    
    def save_image_prompts(self, content_data: Dict):
        """
        保存图片提示词到文件：4 张故事图 + 1 张封面（带短标题）
        """
        prompts_file = os.path.join(self.image_dir, "image_prompts.txt")
        
        with open(prompts_file, 'w', encoding='utf-8') as f:
            f.write("# AI绘画提示词\n\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            # 保存正文内容（用于后续分段叠加到图片上）
            content = content_data.get("content", "").strip()
            if content:
                f.write(f"## 正文内容\n\n")
                f.write(f"{content}\n\n")
                f.write("---\n\n")
            
            # 故事图：至少 4 张
            image_prompts = content_data.get("image_prompts", [])[:4]
            for idx, prompt_data in enumerate(image_prompts, start=1):
                f.write(f"## 图{idx}: {prompt_data.get('scene', '')}\n\n")
                f.write(f"```\n{prompt_data.get('prompt', '')}\n```\n\n")
            
            # 封面：短标题 + 带标题的 prompt
            cover = content_data.get("cover", {})
            if cover.get("title") and cover.get("prompt"):
                f.write(f"## 封面: {cover.get('title', '')}\n\n")
                f.write(f"```\n{cover.get('prompt', '')}\n```\n\n")
        
        print(f"✅ 图片提示词已保存: {prompts_file}")
    
    def save_full_content(self, content_data: Dict, raw_content: str):
        """
        保存完整内容到Markdown文件
        
        Args:
            content_data: 生成的内容数据
            raw_content: 原始输入内容
        """
        md_file = os.path.join(self.image_dir, "content.md")
        
        with open(md_file, 'w', encoding='utf-8') as f:
            f.write("# 小红书文案预览\n\n")
            f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            f.write("## 📕 可选标题\n\n")
            titles = content_data.get("titles", [])
            for idx, title in enumerate(titles, start=1):
                f.write(f"{idx}. {title}\n")
            
            f.write("\n## 📝 正文内容\n\n")
            f.write(content_data.get("content", ""))
            
            f.write("\n\n## 🏷️ 标签\n\n")
            f.write(content_data.get("tags", ""))
            
            f.write("\n\n## 🎨 AI绘画提示词\n\n")
            image_prompts = content_data.get("image_prompts", [])[:4]
            for idx, prompt_data in enumerate(image_prompts, start=1):
                f.write(f"### 图{idx}: {prompt_data.get('scene', '')}\n\n")
                f.write(f"```\n{prompt_data.get('prompt', '')}\n```\n\n")
            cover = content_data.get("cover", {})
            if cover.get("title") and cover.get("prompt"):
                f.write(f"### 封面: {cover.get('title', '')}\n\n")
                f.write(f"```\n{cover.get('prompt', '')}\n```\n\n")
            
            f.write("\n---\n\n")
            f.write("## 📄 原始输入内容\n\n")
            f.write(raw_content)
        
        print(f"✅ 完整内容已保存: {md_file}")
    
    def run(self):
        """运行主流程"""
        try:
            print("=" * 60)
            print("🚀 老北京文化·小红书内容生成器")
            print("=" * 60)
            
            # 1. 读取输入文件
            raw_content = self.read_input_file()
            
            # 2. 生成内容
            print("\n🤖 正在调用AI生成内容...")
            content_data = self.generate_content(raw_content)
            
            # 3. 保存到Excel
            print("\n💾 正在保存到Excel...")
            self.save_to_excel(content_data, raw_content)
            
            # 4. 保存图片提示词
            print("\n💾 正在保存图片提示词...")
            self.save_image_prompts(content_data)
            
            # 5. 保存完整内容
            print("\n💾 正在保存完整内容...")
            self.save_full_content(content_data, raw_content)
            
            print("\n" + "=" * 60)
            print("✅ 所有任务完成！")
            print(f"📁 Excel文件: {self.config['output_excel']}")
            print(f"📁 图片目录: {self.image_dir}")
            print("=" * 60)
            
        except Exception as e:
            print(f"\n❌ 错误: {e}")
            raise


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description="老北京文化·小红书内容生成器")
    parser.add_argument(
        "-c", "--config",
        default="config.json",
        help="配置文件路径 (默认: config.json)"
    )
    
    args = parser.parse_args()
    
    generator = RedBookContentGenerator(config_path=args.config)
    generator.run()


if __name__ == "__main__":
    main()
